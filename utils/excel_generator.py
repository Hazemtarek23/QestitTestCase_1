import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.datavalidation import DataValidation
import os
from datetime import datetime
from typing import List, Dict, Any, Optional


class TestCaseExcelGenerator:
    """Enhanced Excel generator for comprehensive test case documentation"""

    def __init__(self):
        self.header_style = {
            'font': Font(bold=True, color="FFFFFF", size=12),
            'fill': PatternFill(start_color="366092", end_color="366092", fill_type="solid"),
            'alignment': Alignment(horizontal="center", vertical="center", wrap_text=True),
            'border': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        }

        self.cell_style = {
            'alignment': Alignment(wrap_text=True, vertical="top", horizontal="left"),
            'border': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        }

        self.priority_colors = {
            'High': PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid"),
            'Medium': PatternFill(start_color="FFF2E6", end_color="FFF2E6", fill_type="solid"),
            'Low': PatternFill(start_color="E6F7E6", end_color="E6F7E6", fill_type="solid")
        }

    def generate_comprehensive_excel(self, test_cases: List[Dict[str, Any]],
                                     output_path: str,
                                     title: str = "Test Case Documentation",
                                     include_summary: bool = True) -> str:
        """
        Generate comprehensive Excel file with multiple sheets and enhanced formatting

        Args:
            test_cases: List of test case dictionaries
            output_path: Path where Excel file will be saved
            title: Title for the document
            include_summary: Whether to include summary sheet

        Returns:
            str: Success message or error description
        """
        try:
            # Ensure output directory exists
            os.makedirs(os.path.dirname(output_path), exist_ok=True)

            # Create workbook
            wb = Workbook()

            # Remove default sheet if we're adding multiple sheets
            if include_summary:
                wb.remove(wb.active)

            # Generate sheets
            if include_summary:
                self._create_summary_sheet(wb, test_cases, title)

            self._create_detailed_test_cases_sheet(wb, test_cases, title)
            self._create_test_execution_sheet(wb, test_cases)
            self._create_requirements_traceability_sheet(wb, test_cases)

            # Save workbook
            wb.save(output_path)

            return f"Excel file generated successfully at: {output_path}"

        except Exception as e:
            return f"Failed to generate Excel file: {str(e)}"

    def _create_summary_sheet(self, wb: Workbook, test_cases: List[Dict], title: str):
        """Create summary overview sheet"""
        ws = wb.create_sheet("Summary", 0)

        # Title
        ws.merge_cells("A1:F1")
        ws["A1"] = title
        ws["A1"].font = Font(size=16, bold=True)
        ws["A1"].alignment = Alignment(horizontal="center")

        # Generation info
        ws["A3"] = "Generated Date:"
        ws["B3"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws["A4"] = "Total Test Cases:"
        ws["B4"] = len(test_cases)

        # Statistics
        priority_stats = self._calculate_priority_stats(test_cases)
        category_stats = self._calculate_category_stats(test_cases)
        risk_stats = self._calculate_risk_stats(test_cases)

        # Priority breakdown
        ws["A6"] = "Priority Distribution:"
        ws["A6"].font = Font(bold=True)
        row = 7
        for priority, count in priority_stats.items():
            ws[f"A{row}"] = f"  {priority}:"
            ws[f"B{row}"] = count
            ws[f"B{row}"].fill = self.priority_colors.get(priority, PatternFill())
            row += 1

        # Category breakdown
        ws[f"A{row + 1}"] = "Category Distribution:"
        ws[f"A{row + 1}"].font = Font(bold=True)
        row += 2
        for category, count in category_stats.items():
            ws[f"A{row}"] = f"  {category}:"
            ws[f"B{row}"] = count
            row += 1

        # Risk breakdown
        ws[f"A{row + 1}"] = "Risk Level Distribution:"
        ws[f"A{row + 1}"].font = Font(bold=True)
        row += 2
        for risk, count in risk_stats.items():
            ws[f"A{row}"] = f"  {risk}:"
            ws[f"B{row}"] = count
            row += 1

        # Adjust column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15

    def _create_detailed_test_cases_sheet(self, wb: Workbook, test_cases: List[Dict], title: str):
        """Create detailed test cases sheet"""
        if "Summary" in [sheet.title for sheet in wb.worksheets]:
            ws = wb.create_sheet("Detailed Test Cases")
        else:
            ws = wb.active
            ws.title = "Detailed Test Cases"

        # Prepare data
        df_data = []
        for tc in test_cases:
            # Handle complex fields
            preconditions_text = self._format_list_field(tc.get('preconditions', []))
            test_steps_text = self._format_test_steps(tc.get('test_steps', []))
            validation_criteria_text = self._format_list_field(tc.get('validation_criteria', []))
            dependencies_text = self._format_list_field(tc.get('dependencies', []))

            df_data.append({
                'Test ID': tc.get('test_id', ''),
                'Test Name': tc.get('test_name', ''),
                'Module': tc.get('module', ''),
                'Sub Module': tc.get('sub_module', ''),
                'Action': tc.get('action', ''),
                'Description': tc.get('description', ''),
                'Objective': tc.get('objective', ''),
                'Test Type': tc.get('test_type', ''),
                'Category': tc.get('category', ''),
                'Priority': tc.get('priority', ''),
                'Risk Level': tc.get('risk_level', ''),
                'Preconditions': preconditions_text,
                'Test Steps': test_steps_text,
                'Expected Result': tc.get('expected_result', ''),
                'Validation Criteria': validation_criteria_text,
                'Test Data': self._format_test_data(tc.get('test_data', {})),
                'Dependencies': dependencies_text,
                'Execution Time': tc.get('estimated_execution_time', ''),
                'Notes': tc.get('notes', '')
            })

        # Create DataFrame and add to worksheet
        df = pd.DataFrame(df_data)

        # Add headers with formatting
        headers = list(df.columns)
        ws.append(headers)

        # Apply header formatting
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.font = self.header_style['font']
            cell.fill = self.header_style['fill']
            cell.alignment = self.header_style['alignment']
            cell.border = self.header_style['border']

        # Add data rows
        for idx, row in df.iterrows():
            ws.append(list(row))

            # Apply formatting to data rows
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=idx + 2, column=col)
                cell.alignment = self.cell_style['alignment']
                cell.border = self.cell_style['border']

                # Priority-based coloring
                if headers[col - 1] == 'Priority':
                    priority = cell.value
                    if priority in self.priority_colors:
                        cell.fill = self.priority_colors[priority]

        # Set column widths
        column_widths = {
            'A': 12,  # Test ID
            'B': 30,  # Test Name
            'C': 15,  # Module
            'D': 15,  # Sub Module
            'E': 12,  # Action
            'F': 35,  # Description
            'G': 25,  # Objective
            'H': 12,  # Test Type
            'I': 15,  # Category
            'J': 10,  # Priority
            'K': 10,  # Risk Level
            'L': 25,  # Preconditions
            'M': 50,  # Test Steps
            'N': 30,  # Expected Result
            'O': 25,  # Validation Criteria
            'P': 20,  # Test Data
            'Q': 20,  # Dependencies
            'R': 12,  # Execution Time
            'S': 25  # Notes
        }

        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width

    def _create_test_execution_sheet(self, wb: Workbook, test_cases: List[Dict]):
        """Create test execution tracking sheet"""
        ws = wb.create_sheet("Test Execution")

        # Headers for execution tracking
        execution_headers = [
            'Test ID', 'Test Name', 'Priority', 'Category', 'Assigned To',
            'Execution Date', 'Status', 'Actual Result', 'Defects Found',
            'Comments', 'Re-test Required', 'Sign-off'
        ]

        ws.append(execution_headers)

        # Apply header formatting
        for col, header in enumerate(execution_headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.font = self.header_style['font']
            cell.fill = self.header_style['fill']
            cell.alignment = self.header_style['alignment']
            cell.border = self.header_style['border']

        # Add test case data for execution
        for tc in test_cases:
            ws.append([
                tc.get('test_id', ''),
                tc.get('test_name', ''),
                tc.get('priority', ''),
                tc.get('category', ''),
                '',  # Assigned To
                '',  # Execution Date
                'Not Executed',  # Status
                '',  # Actual Result
                '',  # Defects Found
                '',  # Comments
                'No',  # Re-test Required
                ''  # Sign-off
            ])

        # Add data validation for Status column
        status_validation = DataValidation(
            type="list",
            formula1='"Not Executed,Pass,Fail,Blocked,Skip"',
            allow_blank=False
        )
        ws.add_data_validation(status_validation)
        status_validation.add(f'G2:G{len(test_cases) + 1}')

        # Add data validation for Re-test Required column
        retest_validation = DataValidation(
            type="list",
            formula1='"Yes,No"',
            allow_blank=False
        )
        ws.add_data_validation(retest_validation)
        retest_validation.add(f'K2:K{len(test_cases) + 1}')

        # Set column widths
        execution_widths = {
            'A': 12, 'B': 30, 'C': 10, 'D': 15, 'E': 15, 'F': 12,
            'G': 12, 'H': 25, 'I': 15, 'J': 25, 'K': 12, 'L': 15
        }

        for col_letter, width in execution_widths.items():
            ws.column_dimensions[col_letter].width = width

    def _create_requirements_traceability_sheet(self, wb: Workbook, test_cases: List[Dict]):
        """Create requirements traceability matrix"""
        ws = wb.create_sheet("Requirements Traceability")

        # Headers
        traceability_headers = [
            'Requirement ID', 'Requirement Description', 'Test Case ID',
            'Test Case Name', 'Coverage Status', 'Priority', 'Notes'
        ]

        ws.append(traceability_headers)

        # Apply header formatting
        for col, header in enumerate(traceability_headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.font = self.header_style['font']
            cell.fill = self.header_style['fill']
            cell.alignment = self.header_style['alignment']
            cell.border = self.header_style['border']

        # Extract unique modules/features for requirements mapping
        modules = set()
        for tc in test_cases:
            if tc.get('module'):
                modules.add(tc.get('module'))

        # Create requirement entries (this is a template - actual requirements would come from analysis)
        req_counter = 1
        for module in sorted(modules):
            related_tests = [tc for tc in test_cases if tc.get('module') == module]
            for tc in related_tests:
                ws.append([
                    f'REQ-{req_counter:03d}',  # Requirement ID
                    f'{module} - {tc.get("objective", "Functional requirement")}',  # Requirement Description
                    tc.get('test_id', ''),  # Test Case ID
                    tc.get('test_name', ''),  # Test Case Name
                    'Covered',  # Coverage Status
                    tc.get('priority', ''),  # Priority
                    f'Covered by test case {tc.get("test_id", "")}'  # Notes
                ])
                req_counter += 1

        # Set column widths
        traceability_widths = {
            'A': 15, 'B': 40, 'C': 15, 'D': 30, 'E': 12, 'F': 10, 'G': 25
        }

        for col_letter, width in traceability_widths.items():
            ws.column_dimensions[col_letter].width = width

    def _format_list_field(self, field_value) -> str:
        """Format list fields for Excel display"""
        if isinstance(field_value, list):
            return '\n'.join([f"• {item}" for item in field_value])
        elif isinstance(field_value, str):
            return field_value
        else:
            return str(field_value) if field_value else ''

    def _format_test_steps(self, test_steps) -> str:
        """Format test steps with numbering"""
        if isinstance(test_steps, list):
            return '\n'.join([f"{i + 1}. {step}" for i, step in enumerate(test_steps)])
        elif isinstance(test_steps, str):
            return test_steps
        else:
            return str(test_steps) if test_steps else ''

    def _format_test_data(self, test_data) -> str:
        """Format test data dictionary"""
        if isinstance(test_data, dict):
            formatted_data = []
            for key, value in test_data.items():
                formatted_data.append(f"{key}: {value}")
            return '\n'.join(formatted_data)
        else:
            return str(test_data) if test_data else ''

    def _calculate_priority_stats(self, test_cases: List[Dict]) -> Dict[str, int]:
        """Calculate priority distribution statistics"""
        stats = {'High': 0, 'Medium': 0, 'Low': 0}
        for tc in test_cases:
            priority = tc.get('priority', 'Medium')
            if priority in stats:
                stats[priority] += 1
        return stats

    def _calculate_category_stats(self, test_cases: List[Dict]) -> Dict[str, int]:
        """Calculate category distribution statistics"""
        stats = {}
        for tc in test_cases:
            category = tc.get('category', 'Unknown')
            stats[category] = stats.get(category, 0) + 1
        return stats

    def _calculate_risk_stats(self, test_cases: List[Dict]) -> Dict[str, int]:
        """Calculate risk level distribution statistics"""
        stats = {'High': 0, 'Medium': 0, 'Low': 0}
        for tc in test_cases:
            risk = tc.get('risk_level', 'Medium')
            if risk in stats:
                stats[risk] += 1
        return stats


# Convenience functions for backward compatibility and easy usage
def generate_test_case_excel(test_cases: List[Dict[str, Any]], output_path: str,
                             title: str = "Test Case Documentation") -> str:
    """
    Generate comprehensive Excel file from test cases (enhanced version)

    Args:
        test_cases: List of test case dictionaries
        output_path: Path where Excel file will be saved
        title: Title for the document

    Returns:
        str: Success message or error description
    """
    generator = TestCaseExcelGenerator()
    return generator.generate_comprehensive_excel(test_cases, output_path, title)


def generate_simple_excel_report(test_cases: List[Dict[str, Any]], output_path: str,
                                 title: str = "Test Cases") -> str:
    """
    Generate simple Excel report (backward compatibility)

    Args:
        test_cases: List of test case dictionaries
        output_path: Path where Excel file will be saved
        title: Title for the document

    Returns:
        str: Success message or error description
    """
    try:
        os.makedirs(os.path.dirname(output_path), exist_ok=True)

        wb = Workbook()
        ws = wb.active
        ws.title = "Test Cases"

        # Title
        ws.merge_cells("A1:E1")
        ws["A1"] = title
        ws["A1"].font = Font(size=14, bold=True)
        ws["A1"].alignment = Alignment(horizontal="center")

        # Headers
        headers = ["Test Case ID", "Test Name", "Description", "Priority", "Status"]
        ws.append(headers)

        # Format headers
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=2, column=col)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        # Add data
        for case in test_cases:
            ws.append([
                case.get("test_id", ""),
                case.get("test_name", ""),
                case.get("description", ""),
                case.get("priority", "Medium"),
                case.get("status", "Not Run")
            ])

        # Set column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 40
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 15

        wb.save(output_path)
        return f"Simple Excel report generated successfully at: {output_path}"

    except Exception as e:
        return f"Error generating simple Excel report: {str(e)}"


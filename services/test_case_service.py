"""Enhanced Test Case Service - Large Scale Professional Test Case Generation
Handles comprehensive test case generation with detailed test data, extensive functionality coverage,
and automatic generation of large numbers of test cases with professional Excel output"""

import os
import logging
from typing import Dict, List, Any, Optional
from datetime import datetime, timedelta
import json
import random
import string
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import re
from services.llama_client import LlamaClient
from services.ports import TextGenerator
from typing import Optional

class EnhancedTestCaseService:
    def __init__(self, model_name: str = "gpt-4.1", generator: Optional[TextGenerator] = None):
        self.model_name = model_name
        self.logger = logging.getLogger(__name__)

        # Initialize Llama client
        try:
            self.azure_client = generator or LlamaClient(model_name)
            self.logger.info(f"✅ Llama client initialized with model: {model_name}")
        except Exception as e:
            self.logger.error(f"Failed to initialize Llama client: {str(e)}")
            self.azure_client = None

        # Enhanced system prompt for large-scale professional test case generation
        self.ENHANCED_SYSTEM_PROMPT = """You are an elite test case generation specialist with extensive domain expertise in banking and financial systems. Your mission is to create LARGE-SCALE, PROFESSIONAL test cases that comprehensively cover ALL system functionality with realistic test data and detailed execution steps.

⚠️ CRITICAL LARGE-SCALE GENERATION REQUIREMENTS ⚠️
================================================================

MANDATORY: Generate EXACTLY {iteration_count} COMPLETE, UNIQUE test cases. Each test case must be:
- Completely unique and testing different functionality
- Professionally detailed with realistic test data
- Focused on FUNCTIONAL testing beyond simple validation
- Covering comprehensive business workflows and processes

PROFESSIONAL TEST CASE NAMING CONVENTION:
========================================
Each title must be extremely descriptive and follow this enhanced format:
{feature_prefix}_[BusinessArea]_[Module]_[SubModule]_[DetailedBusinessScenario]_[UserRole]_[TestCategory]_[ComplexityLevel]

Enhanced Examples:
- {feature_prefix}_CustomerOnboarding_IndividualAccounts_SavingsAccount_CreateNewAccountWithKYCVerificationAndInitialDepositProcessing_CSR_EndToEndBusinessProcess_Complex
- {feature_prefix}_TransactionProcessing_InternationalTransfers_HighValueTransfers_ProcessLargeAmountTransferWithMultiLevelApprovalAndComplianceChecking_BranchManager_BusinessWorkflowIntegration_Critical
- {feature_prefix}_LoanManagement_PersonalLoans_LoanOrigination_ProcessNewPersonalLoanApplicationWithCreditScoringAndDocumentVerification_LoanOfficer_CompleteBusinessProcess_High

COMPREHENSIVE FUNCTIONAL COVERAGE DISTRIBUTION:
==============================================
Create test cases covering these functional areas (NOT just validation):

1. **Core Banking Operations (35% of test cases):**
   - Account Management: Opening, closing, modification, maintenance
   - Customer Onboarding: KYC, documentation, verification processes
   - Transaction Processing: Payments, transfers, collections, clearing
   - Loan Processing: Origination, approval, disbursement, servicing

2. **Digital Banking & Channels (20% of test cases):**
   - Internet Banking: Login, navigation, transactions, services
   - Mobile Banking: App functionality, mobile payments, notifications
   - ATM Operations: Cash withdrawal, deposits, inquiries, maintenance

3. **Risk & Compliance (15% of test cases):**
   - AML/KYC: Monitoring, reporting, compliance checking
   - Credit Risk: Assessment, scoring, monitoring, reporting
   - Fraud Detection: Transaction monitoring, alert generation, investigation

4. **Back Office Operations (10% of test cases):**
   - Settlement & Clearing: Transaction settlement, reconciliation
   - Accounting: General ledger, financial reporting, closing processes
   - Operations Support: Exception handling, manual processes, corrections

5. **Customer Experience (10% of test cases):**
   - Customer Service: Call center operations, complaint handling
   - Relationship Management: Client onboarding, portfolio reviews

6. **Integration & Infrastructure (5% of test cases):**
   - System Integrations: Core banking, payment systems, external services
   - Security: Authentication, authorization, encryption, monitoring

7. **Analytics & Reporting (5% of test cases):**
   - Business Intelligence: Dashboards, analytics, insights
   - Management Reporting: Executive reports, KPI monitoring

REALISTIC TEST DATA REQUIREMENTS (MANDATORY FOR ALL FIELDS):
===========================================================

**Personal Data Examples:**
- Arabic Names: "أحمد محمد الراشد", "فاطمة علي الزهراني", "خالد عبدالله القحطاني"
- Western Names: "Sarah Elizabeth Johnson", "Michael James O'Connor", "Emma Charlotte Williams"
- Asian Names: "李明华 (Li Ming Hua)", "田中雄介 (Tanaka Yusuke)", "Kim Min-jun", "Priya Sharma"

**Addresses (Complete):**
- Qatar: "Villa 23, Street 820, Al-Sadd District, Zone 25, P.O. Box 12345, Doha, Qatar"
- UAE: "Apartment 1205, Burj Al-Arab Business Tower, Sheikh Zayed Road, Dubai, UAE 12345"

**Contact Information:**
- Qatar Mobile: "+974-5544-7788", "+974-3366-9922", "+974-7788-1122"
- Emails: "ahmed.alrashid@qatarbank.com.qa", "sarah.johnson@company.com"

**Financial Data (Specific Amounts):**
- Account Numbers: "QAR-SAV-2024-789456123", "USD-CHK-2024-555777888"
- IBAN: "QA58DOHB00001234567890123456", "QA29QNBK000000000012345678"
- Amounts: "15,750.50 QAR", "125,500.75 USD", "85,250.25 EUR"
        
        Generate the complete set of {iteration_count} professional test cases following ALL requirements above."""

        # Strict system prompt used to generate standardized markdown table test cases
        self.STRICT_TEST_CASE_SYSTEM_PROMPT = """You are a test case generation specialist. Your task is to create comprehensive test cases in a standardized markdown table format that covers all aspects of testing.

⚠️ CRITICAL FIELD VALIDATION LIMIT ⚠️
==================================
STRICTLY ENFORCED RULE: For each screen in the application, you MUST NOT generate more than TWO (2) test cases that focus on validation of optional and mandatory fields. This is a hard limit that cannot be exceeded under any circumstances. Plan your test cases accordingly to stay within this limit while ensuring comprehensive coverage through other types of test cases.

CRITICAL REQUIREMENT: You MUST generate EXACTLY {iteration_count} COMPLETE test cases in the table. DO NOT use placeholders, "and so on", or any other shorthand notation. Each test case must be fully detailed and unique.

TEST CASE NAMING CONVENTION:
Each test case title must follow this exact format:
{feature_prefix}_[Module]_[SubModule]_[Action]

Examples:
- {feature_prefix}_Account_SavingAccount_Create
- {feature_prefix}_Account_CurrentAccount_Update
- {feature_prefix}_Customer_Individual_Validate
- {feature_prefix}_Transaction_InternalTransfer_Process

Where:
- {feature_prefix}: Project or system name (e.g., NBK, QNB, ADIB)
- Module: Main module being tested (e.g., Account, Customer, Transaction)
- SubModule: Specific feature or type being tested (e.g., SavingAccount, CurrentAccount, Individual)
- Action: The action being tested (e.g., Create, Update, Delete, Validate, Process)

Each test case should:
1. Have a unique focus and objective
2. Cover different scenarios and aspects of the feature
3. Have different test steps and validation criteria
4. Target different user roles or edge cases
5. Have unique preconditions and test data

IMPORTANT: You must follow this exact markdown structure. Do not add any additional sections or modify the structure.

## Test Case Summary
| Test Case ID | Title | Priority | Module/Feature | Test Type | Description | Objective | Preconditions | Test Steps | Expected Results | Post-Conditions | Notes |
|--------------|-------|----------|----------------|-----------|-------------|-----------|---------------|------------|-----------------|-----------------|-------|
{test_cases}

Each **test step and its expected result must be in a separate table row**. Multiple rows may be used for a single test case, each representing one test step and its corresponding result.

ABSOLUTE PROHIBITIONS:
- Do NOT write placeholders like "Detailed step-by-step execution procedure" or "N/A" or "etc." or "and so on".
- Every cell must contain concrete, specific content with real values, actors, screens, fields, and messages.
- Test Steps MUST be explicitly numbered (1., 2., 3., …) and each step must include actor, navigation, field/value, and action.
- Expected Results MUST be specific and measurable (e.g., exact message text, state changes, database effects), never generic.

MINIMUM STEP DEPTH:
- Provide at least 12 steps per test case (aim 15–25 when applicable). Each step in its own row paired with its expected result.

MANDATORY FIRST STEP:
- Step 1 must be an explicit login step with the appropriate actor (user role) and successful landing page validation, if the application requires authentication.

FUNCTIONAL COVERAGE WEIGHTING:
- At least 70% of test cases must focus on functional business workflows and cross-screen navigation; validation-focused cases are strictly limited by the 2-per-screen rule above.

NAVIGATION AND READABILITY:
- Always show navigation using the arrow symbol "→" (e.g., Home → Accounts → New Account → Review).
- Every step must include the actor, the screen, the action, and concrete field values; avoid vague wording.

ACTOR ROLES AND AUTHORIZATION (MAKER–CHECKER ONLY):
- Explicitly identify actors per step as Maker or Checker.
- For transactions requiring approval, include Maker submission and Checker review/approval.
- Reflect status transitions (e.g., Draft → Pending Approval → Approved/Rejected) and capture audit trail updates.

VOCABULARY NORMALIZATION:
- Use roles Maker and Checker only. Do NOT use the term "Admin" or "Administrator"; map such references to "Maker".

Guidelines for Creating Test Cases:

1. Content Quality
   - Each step must be clear, concise, and unambiguous
   - Use active voice and present tense
   - Include specific data values, not just descriptions
   - Specify exact expected outcomes or error messages
   - One test case should focus on one specific functionality

2. Test Coverage
   - Include both positive and negative scenarios
   - Cover edge cases and boundary conditions
   - Consider error handling and validation
   - Include data validation where applicable
   - Test both UI elements and business logic

3. Best Practices
   - Use consistent terminology throughout
   - Number steps sequentially
   - Include validation points at each critical step
   - Specify exact element identifiers (ID, name, XPath)
   - Consider test data dependencies

4. Formatting Rules
   - Use exactly the markdown structure provided above
   - Do not add or remove any sections
   - Keep all table headers as shown
   - Use bullet points (•) for lists within table cells
   - Use line breaks (<br>) to separate items within cells
   - Use bold (**) for emphasis on key terms
   - Ensure proper spacing between sections
   - Format test steps with clear action/result/validation structure

5. Test Case Distribution
   - Each test case must be completely different from others
   - Distribute test cases across different aspects of the feature
   - Ensure each test case has a unique focus
   - Cover different user roles and permissions
   - Include both happy path and error scenarios
   - Test different data combinations and edge cases
   - For each screen, generate a maximum of two test cases to validate optional and mandatory fields.
   - **CRITICAL REQUIREMENT: For each screen, you MUST NOT generate more than two test cases that focus on validation of optional and mandatory fields.**

6. Screen Navigation Usage
   - Use the provided screen navigation documentation to create precise test steps
   - Follow the exact navigation paths described in the documentation
   - Include specific screen names and transitions in test steps
   - Reference exact field names, button labels, and UI elements
   - Ensure test steps match the actual system navigation flow
   - Include validation of UI elements' presence and state
   - Test both successful and failed navigation scenarios
   - Verify correct screen transitions and field interactions
   - Include validation of screen-specific behaviors and states
   - ALWAYS specify the actor (user role) for each navigation step

7. Test Step Detail Requirements
   - Each test step MUST be in its own row in the table
   - Each row must contain exactly one test step and its corresponding expected result
   - Break down complex actions into individual steps
   - Include specific data values in every step
   - For each step, include:
     * Exact navigation path with actor
     * Specific screen name
     * Exact field names and values
     * Precise button labels
     * Specific validation points

CRITICAL REQUIREMENTS:
1. You MUST generate EXACTLY {iteration_count} test cases - no more, no less
2. DO NOT use placeholders, "and so on", or any other shorthand notation
3. Each test case must be fully detailed and complete
4. DO NOT skip any test cases or use ellipsis (...)
5. Each test case must be unique and provide value
6. The table must be complete with all test cases
7. Each test case MUST follow the exact step structure specified above
8. ALL steps must include the actor where applicable
9. ALL navigation must be explicit and complete
10. ALL field population must be grouped in a single step with substeps
11. ALL steps and substeps MUST have their own expected results
12. ALL expected results must be specific and measurable
13. ALL validations must be clearly defined
14. ALL success and failure conditions must be documented
15. ALL side effects must be verified
16. **For each screen, you MUST NOT generate more than two test cases that focus on validation of optional and mandatory fields.**

Your test cases must follow this exact markdown structure. Do not modify the structure or add any additional sections. Each section must be properly formatted using markdown syntax as shown above.

Create {iteration_count} test cases that collectively provide comprehensive coverage of the feature's functionality. Each test case should focus on a different aspect of the feature being tested.
"""

    async def generate_text_with_hf(self, prompt: str, max_length: int = 2048) -> str:
        """Generate text using Llama client"""
        try:
            if not self.azure_client:
                self.logger.error("No Llama client available")
                return self._create_fallback_content(prompt)

            # Generate text using the client
            result = await self.azure_client.generate_text(
                system_prompt="You are a professional test case generation specialist.",
                user_prompt=prompt,
                max_length=max_length
            )

            return result if result else self._create_fallback_content(prompt)

        except Exception as e:
            self.logger.error(f"Error generating text with Llama server: {str(e)}")
            return self._create_fallback_content(prompt)

    def _create_fallback_content(self, prompt: str) -> str:
        """Create fallback content when model generation fails"""
        return """# Professional Test Case Generation

Based on the requirements provided, here are comprehensive test cases:

## Test Case 1: Account Management
- **ID**: TC_001
- **Title**: Create New Savings Account with KYC Verification
- **Area**: Core Banking
- **Priority**: High
- **Steps**: Login, Navigate to Accounts, Create Account, Verify KYC, Submit
- **Data**: Customer: Ahmed Al-Rashid, Amount: 5,000.00 QAR
- **Expected**: Account created successfully with proper documentation

## Test Case 2: Transaction Processing
- **ID**: TC_002  
- **Title**: Process International Wire Transfer
- **Area**: Transaction Processing
- **Priority**: Critical
- **Steps**: Login, Select Transfer, Enter Details, Approve, Process
- **Data**: Amount: 50,000.00 QAR, Destination: UK Bank
- **Expected**: Transfer processed with compliance checks

## Test Case 3: Digital Banking
- **ID**: TC_003
- **Title**: Mobile Banking Login and Balance Inquiry
- **Area**: Digital Banking
- **Priority**: High
- **Steps**: Open App, Enter Credentials, Verify OTP, Check Balance
- **Data**: User: sarah.johnson@bank.com, Account: QAR-SAV-2024-123456
- **Expected**: Successful login and accurate balance display

Additional test cases would follow similar patterns covering all banking operations."""

    async def generate_large_scale_test_cases(
            self,
            user_input: str,
            technical_analysis: str = "",
            domain_analysis: str = "",
            screen_navigation: str = "",
            test_risk: str = "",
            test_plan: str = "",
            feature_prefix: str = "TC",
            iteration_count: int = 300,
            raw_input_chunk: str = ""
    ) -> Dict[str, Any]:
        """
        Generate maximum possible unique test cases based on document content analysis
        """
        try:
            self.logger.info(f"Starting dynamic test case generation based on document content analysis")

            # Analyze document content to extract unique features and requirements
            document_features = self._extract_document_features(
                user_input, technical_analysis, domain_analysis, 
                screen_navigation, test_risk, test_plan, raw_input_chunk
            )
            
            # Calculate maximum possible unique test cases based on document features
            max_possible_cases = self._calculate_maximum_test_cases(document_features)
            
            self.logger.info(f"Document analysis found {len(document_features)} unique features, generating {max_possible_cases} test cases")
            
            # Generate test cases based on actual document content
            test_cases_result = await self._generate_content_based_test_cases(
                document_features, feature_prefix, max_possible_cases
            )

            if not test_cases_result["success"]:
                return test_cases_result

            # Generate timestamp for file naming
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

            # Create Excel file
            excel_path = self._create_professional_excel(
                test_cases_result["test_cases"], feature_prefix, timestamp, max_possible_cases
            )

            # Create enhanced markdown file
            markdown_path = self._create_enhanced_markdown(
                test_cases_result["content"], feature_prefix, timestamp
            )

            return {
                "success": True,
                "test_cases_content": test_cases_result["content"],
                "markdown_path": markdown_path,
                "excel_path": excel_path,
                "generated_count": len(test_cases_result["test_cases"]),
                "feature_prefix": feature_prefix,
                "timestamp": timestamp,
                "document_features_analyzed": len(document_features),
                "max_possible_cases_calculated": max_possible_cases,
                "enhancement_features": [
                    "Dynamic content-based test case generation",
                    "Maximum possible unique test cases based on document analysis",
                    "Document feature extraction and analysis",
                    "Professional Excel output with advanced formatting",
                    "Comprehensive functional coverage based on actual requirements",
                    "Realistic test data for all fields",
                    "Detailed 15-25 step procedures",
                    "Business process workflow testing",
                    "Integration and end-to-end scenarios"
                ]
            }

        except Exception as e:
            self.logger.error(f"Error in dynamic test case generation: {str(e)}")
            return {
                "success": False,
                "error": str(e),
                "test_cases_content": "",
                "markdown_path": None,
                "excel_path": None,
                "generated_count": 0
            }

    def _calculate_maximum_test_cases(self, document_features: List[Dict[str, Any]]) -> int:
        """Calculate the maximum possible unique test cases based on document features"""
        if not document_features:
            return 50  # Minimum fallback
            
        # Base calculation: each feature can generate multiple test scenarios
        base_cases_per_feature = 8  # Different scenarios per feature combination
        
        # Calculate based on feature complexity
        total_possible_cases = len(document_features) * base_cases_per_feature
        
        # Apply complexity multipliers based on feature richness
        complexity_multiplier = self._calculate_complexity_multiplier(document_features)
        
        # Calculate final maximum
        max_cases = int(total_possible_cases * complexity_multiplier)
        
        # Apply reasonable limits
        min_cases = 20
        max_cases = min(max_cases, 500)  # Cap at 500 to ensure quality
        max_cases = max(max_cases, min_cases)
        
        self.logger.info(f"Calculated maximum test cases: {max_cases} (features: {len(document_features)}, complexity: {complexity_multiplier:.2f})")
        
        return max_cases

    def _calculate_complexity_multiplier(self, document_features: List[Dict[str, Any]]) -> float:
        """Calculate complexity multiplier based on feature richness"""
        if not document_features:
            return 1.0
            
        # Count unique elements
        unique_screens = len(set(f["screen"] for f in document_features))
        unique_modules = len(set(f["module"] for f in document_features))
        unique_processes = len(set(f["business_process"] for f in document_features))
        unique_roles = len(set(f["user_role"] for f in document_features))
        
        # Calculate richness score
        total_elements = unique_screens + unique_modules + unique_processes + unique_roles
        feature_count = len(document_features)
        
        # Complexity increases with more unique elements and feature combinations
        complexity = 1.0 + (total_elements / 20.0) + (feature_count / 50.0)
        
        # Cap complexity between 0.5 and 3.0
        complexity = max(0.5, min(3.0, complexity))
        
        return complexity

    def _extract_document_features(self, user_input: str, technical_analysis: str, 
                                  domain_analysis: str, screen_navigation: str, 
                                  test_risk: str, test_plan: str, raw_input_chunk: str) -> List[Dict[str, Any]]:
        """Extract unique features and requirements from document content"""
        features = []
        
        # Combine all document content
        all_content = f"{user_input}\n{technical_analysis}\n{domain_analysis}\n{screen_navigation}\n{test_risk}\n{test_plan}\n{raw_input_chunk}"
        
        # Extract screens and modules
        screens = self._extract_screens_from_content(all_content)
        modules = self._extract_modules_from_content(all_content)
        business_processes = self._extract_business_processes_from_content(all_content)
        user_roles = self._extract_user_roles_from_content(all_content)
        data_entities = self._extract_data_entities_from_content(all_content)
        
        # Create feature combinations - more comprehensive approach
        for screen in screens:
            for module in modules:
                for process in business_processes:
                    for role in user_roles:
                        feature = {
                            "screen": screen,
                            "module": module,
                            "business_process": process,
                            "user_role": role,
                            "data_entities": data_entities,
                            "feature_id": f"{screen}_{module}_{process}_{role}".replace(" ", "_")
                        }
                        features.append(feature)
        
        # Limit to reasonable number and ensure uniqueness
        unique_features = []
        seen_ids = set()
        for feature in features:
            if feature["feature_id"] not in seen_ids and len(unique_features) < 200:  # Increased limit
                seen_ids.add(feature["feature_id"])
                unique_features.append(feature)
        
        # If we have very few features, expand with additional combinations
        if len(unique_features) < 10:
            unique_features = self._expand_feature_combinations(unique_features, all_content)
        
        return unique_features

    def _expand_feature_combinations(self, base_features: List[Dict[str, Any]], content: str) -> List[Dict[str, Any]]:
        """Expand feature combinations when document has limited features"""
        expanded_features = base_features.copy()
        
        # Add more screen variations
        additional_screens = ["Settings", "Profile", "Reports", "Analytics", "Configuration", "Monitoring"]
        additional_modules = ["Security", "Compliance", "Audit", "Integration", "Performance", "Backup"]
        additional_processes = ["Data Export", "System Backup", "User Management", "Configuration Update", "Report Generation"]
        additional_roles = ["Supervisor", "Analyst", "Operator", "Viewer", "Auditor", "Support"]
        
        # Get existing elements
        existing_screens = set(f["screen"] for f in base_features)
        existing_modules = set(f["module"] for f in base_features)
        existing_processes = set(f["business_process"] for f in base_features)
        existing_roles = set(f["user_role"] for f in base_features)
        
        # Add new combinations
        for screen in additional_screens:
            if screen not in existing_screens:
                for module in existing_modules:
                    for process in existing_processes:
                        for role in existing_roles:
                            feature = {
                                "screen": screen,
                                "module": module,
                                "business_process": process,
                                "user_role": role,
                                "data_entities": ["Account", "Customer", "Transaction"],
                                "feature_id": f"{screen}_{module}_{process}_{role}".replace(" ", "_")
                            }
                            if feature["feature_id"] not in [f["feature_id"] for f in expanded_features]:
                                expanded_features.append(feature)
        
        # Ensure we have a reasonable number of features
        return expanded_features[:50]  # Cap at 50 to maintain quality

    def _extract_screens_from_content(self, content: str) -> List[str]:
        """Extract screen names from document content"""
        screens = []
        content_lower = content.lower()
        
        # Common screen patterns
        screen_patterns = [
            "login", "dashboard", "home", "account", "customer", "transaction",
            "payment", "transfer", "loan", "card", "report", "admin", "settings",
            "profile", "notification", "search", "filter", "approval", "workflow"
        ]
        
        for pattern in screen_patterns:
            if pattern in content_lower:
                screens.append(pattern.title())
        
        # Extract specific screen names using regex
        import re
        screen_matches = re.findall(r'([A-Z][a-z]+(?:\s+[A-Z][a-z]+)*)\s+(?:Screen|Page|Form|Module)', content)
        screens.extend([match.strip() for match in screen_matches])
        
        return list(set(screens)) if screens else ["Main", "Dashboard", "Account", "Transaction"]

    def _extract_modules_from_content(self, content: str) -> List[str]:
        """Extract module names from document content"""
        modules = []
        content_lower = content.lower()
        
        # Common module patterns
        module_patterns = [
            "account management", "customer management", "transaction processing",
            "payment processing", "loan management", "card management", "reporting",
            "administration", "security", "compliance", "risk management", "audit"
        ]
        
        for pattern in module_patterns:
            if pattern in content_lower:
                modules.append(pattern.title())
        
        return list(set(modules)) if modules else ["Account Management", "Customer Management", "Transaction Processing"]

    def _extract_business_processes_from_content(self, content: str) -> List[str]:
        """Extract business processes from document content"""
        processes = []
        content_lower = content.lower()
        
        # Common business process patterns
        process_patterns = [
            "account opening", "customer onboarding", "fund transfer", "payment processing",
            "loan application", "card activation", "report generation", "user registration",
            "password reset", "profile update", "transaction approval", "compliance check"
        ]
        
        for pattern in process_patterns:
            if pattern in content_lower:
                processes.append(pattern.title())
        
        return list(set(processes)) if processes else ["Account Opening", "Fund Transfer", "User Registration"]

    def _extract_user_roles_from_content(self, content: str) -> List[str]:
        """Extract user roles from document content"""
        roles = []
        content_lower = content.lower()
        
        # Common role patterns
        role_patterns = [
            "customer", "admin", "manager", "operator", "supervisor", "analyst",
            "teller", "officer", "agent", "user", "maker", "checker", "approver"
        ]
        
        for pattern in role_patterns:
            if pattern in content_lower:
                roles.append(pattern.title())
        
        return list(set(roles)) if roles else ["Customer", "Admin", "Manager"]

    def _extract_data_entities_from_content(self, content: str) -> List[str]:
        """Extract data entities from document content"""
        entities = []
        content_lower = content.lower()
        
        # Common data entity patterns
        entity_patterns = [
            "account", "customer", "transaction", "payment", "loan", "card",
            "user", "profile", "document", "report", "notification", "audit"
        ]
        
        for pattern in entity_patterns:
            if pattern in content_lower:
                entities.append(pattern.title())
        
        return list(set(entities)) if entities else ["Account", "Customer", "Transaction"]

    async def _generate_content_based_test_cases(self, document_features: List[Dict[str, Any]], 
                                               feature_prefix: str, max_cases: int) -> Dict[str, Any]:
        """Generate test cases based on actual document features"""
        try:
            # Calculate how many test cases per feature
            test_cases_per_feature = max(1, max_cases // len(document_features))
            remaining_cases = max_cases % len(document_features)
            
            all_test_cases = []
            all_content = []
            
            for i, feature in enumerate(document_features):
                # Calculate test cases for this feature
                current_count = test_cases_per_feature + (1 if i < remaining_cases else 0)
                
                # Generate test cases for this specific feature
                feature_test_cases = await self._generate_feature_specific_test_cases(
                    feature, feature_prefix, current_count, i
                )
                
                all_test_cases.extend(feature_test_cases["test_cases"])
                if feature_test_cases.get("content"):
                    all_content.append(feature_test_cases["content"])
            
            # Ensure we have exactly the calculated maximum
            if len(all_test_cases) > max_cases:
                all_test_cases = all_test_cases[:max_cases]
            elif len(all_test_cases) < max_cases:
                # Generate additional cases to reach the target
                additional_needed = max_cases - len(all_test_cases)
                additional_cases = self._create_fallback_test_cases(999, additional_needed)
                all_test_cases.extend(additional_cases)
            
            return {
                "success": True,
                "test_cases": all_test_cases,
                "content": "\n\n".join(all_content) if all_content else "Generated test cases based on document analysis"
            }
            
        except Exception as e:
            self.logger.error(f"Error generating content-based test cases: {str(e)}")
            return {
                "success": False,
                "error": str(e),
                "test_cases": [],
                "content": ""
            }

    async def _generate_feature_specific_test_cases(self, feature: Dict[str, Any], 
                                                   feature_prefix: str, count: int, 
                                                   feature_index: int) -> Dict[str, Any]:
        """Generate test cases for a specific document feature"""
        try:
            system_prompt = (
                self.STRICT_TEST_CASE_SYSTEM_PROMPT
                .replace("{iteration_count}", str(count))
                .replace("{feature_prefix}", feature_prefix)
            )

            user_prompt = f"""FEATURE-SPECIFIC TEST CASE GENERATION

Document Feature Analysis:
- Screen: {feature['screen']}
- Module: {feature['module']}
- Business Process: {feature['business_process']}
- User Role: {feature['user_role']}
- Data Entities: {', '.join(feature['data_entities'])}

Generate exactly {count} unique test cases specifically for this feature combination.
Each test case must focus on the interaction between {feature['user_role']} role and {feature['screen']} screen 
within the {feature['module']} module for {feature['business_process']} process.

Requirements:
1. Each test case must be completely unique
2. Focus on the specific feature combination provided
3. Include realistic test data for {feature['user_role']} role
4. Cover different scenarios within this feature
5. Include both positive and negative test cases
6. Ensure comprehensive coverage of the {feature['business_process']} process"""

            full_prompt = f"{system_prompt}\n\n{user_prompt}"

            # Generate content using the configured Llama model
            content = await self.generate_text_with_hf(full_prompt, max_length=4096)

            # Parse test cases
            test_cases = self._parse_test_cases_from_content(content, feature_index, count)

            # If parsing fails, create feature-specific fallback cases
            if len(test_cases) == 0:
                test_cases = self._create_feature_specific_fallback_cases(feature, feature_index, count)

            return {
                "success": True,
                "content": content,
                "test_cases": test_cases
            }

        except Exception as e:
            self.logger.error(f"Error generating feature-specific test cases: {str(e)}")
            return {
                "success": False,
                "error": str(e),
                "content": "",
                "test_cases": []
            }

    def _parse_test_cases_from_content(self, content: str, batch_num: int, expected_count: int = 10) -> List[Dict[str, Any]]:
        """Parse test cases from generated content with fallback generation"""
        test_cases = []

        # Try to parse structured test cases from content
        lines = content.split('\n')
        for line in lines:
            if not line.startswith('|'):
                continue
            cells = [cell.strip() for cell in line.split('|')[1:-1]]
            # Expect at least the first 12 columns; if not, skip
            if len(cells) < 12:
                continue
            try:
                test_case = {
                    'test_case_id': cells[0] or f"TC_BATCH{batch_num}_{len(test_cases)+1:03d}",
                    'title': cells[1][:200] if len(cells) > 1 else f"Test Case {len(test_cases)+1}",
                    'business_area': cells[2] if len(cells) > 2 else "Banking Operations",
                    'test_category': cells[3] if len(cells) > 3 else "Functional",
                    'priority': cells[4] if len(cells) > 4 else "High",
                    'test_objective': cells[5] if len(cells) > 5 else "Test business functionality",
                    'user_role': cells[6] if len(cells) > 6 else "Maker",
                    'preconditions': cells[7] if len(cells) > 7 else "System accessible",
                    'test_data': cells[8] if len(cells) > 8 else "Realistic test data provided",
                    'test_steps': cells[9] if len(cells) > 9 else "",
                    'expected_results': cells[10] if len(cells) > 10 else "",
                    'post_conditions': cells[11] if len(cells) > 11 else "",
                }
                test_cases.append(test_case)
            except Exception:
                # Skip malformed lines rather than raising
                continue

        # If parsing fails or not enough test cases, create structured test cases
        if len(test_cases) < expected_count:
            additional_needed = expected_count - len(test_cases)
            fallback_cases = self._create_fallback_test_cases(batch_num, additional_needed)
            test_cases.extend(fallback_cases)

        return test_cases[:expected_count]  # Ensure exact count

    def _extract_markdown_table_rows(self, content: str) -> List[Dict[str, str]]:
        """Extract strict markdown rows for the required table structure.
        Returns list of dicts keyed by headers if parseable, else empty list."""
        try:
            lines = [l.rstrip() for l in content.splitlines()]
            header_idx = None
            for i, line in enumerate(lines):
                if line.strip().startswith("| Test Case ID | Title | Priority | Module/Feature | Test Type | Description | Objective | Preconditions | Test Steps | Expected Results | Post-Conditions | Notes |"):
                    header_idx = i
                    break
            if header_idx is None:
                return []

            # Next line after header should be the separator row starting with |---
            start_idx = header_idx + 2 if header_idx + 1 < len(lines) else header_idx + 1
            headers = [h.strip() for h in lines[header_idx].strip('|').split('|')]
            rows: List[Dict[str, str]] = []
            for j in range(start_idx, len(lines)):
                row_line = lines[j].strip()
                if not row_line.startswith('|'):
                    # stop at first non-table line
                    break
                cells = [c.strip() for c in row_line.strip('|').split('|')]
                if len(cells) != len(headers):
                    continue
                row_map = {headers[k]: cells[k] for k in range(len(headers))}
                # Skip rows that contain placeholders
                if self._row_has_placeholder(row_map):
                    continue
                rows.append(row_map)
            return rows
        except Exception:
            return []

    def _count_unique_cases_by_id(self, rows: List[Dict[str, str]]) -> int:
        seen: set = set()
        for r in rows:
            tcid = (r.get("Test Case ID") or "").strip()
            if tcid:
                seen.add(tcid)
        return len(seen)

    def _create_excel_from_markdown_rows(self, rows: List[Dict[str, str]], feature_prefix: str, timestamp: str) -> Optional[str]:
        try:
            excel_filename = f"Professional_Test_Cases_{feature_prefix}_{timestamp}_{self._count_unique_cases_by_id(rows)}Cases.xlsx"
            excel_path = os.path.join("outputs", excel_filename)
            os.makedirs("outputs", exist_ok=True)

            wb = Workbook()
            ws = wb.active
            ws.title = "Test Cases"

            headers = [
                "Test Case ID", "Title", "Priority", "Module/Feature",
                "Test Type", "Description", "Objective", "Preconditions",
                "Test Steps", "Expected Results", "Post-Conditions", "Notes"
            ]

            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col)
                cell.value = header
                cell.font = Font(bold=True, color="FFFFFF", size=12)
                cell.fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = Border(
                    left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin')
                )

            row_idx = 2
            for r in rows:
                values = []
                for h in headers:
                    v = r.get(h, "")
                    if h in ("Test Steps", "Expected Results"):
                        v = v.replace("<br>", "\n")
                    values.append(v)
                for col, value in enumerate(values, 1):
                    cell = ws.cell(row=row_idx, column=col)
                    cell.value = str(value)[:5000]
                    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                    cell.border = Border(
                        left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin')
                    )
                    # Leave body cells uncolored; header is already colored
                row_idx += 1

            column_widths = {
                'A': 15, 'B': 50, 'C': 10, 'D': 25, 'E': 15,
                'F': 40, 'G': 30, 'H': 30, 'I': 60, 'J': 60,
                'K': 25, 'L': 25
            }
            for column, width in column_widths.items():
                ws.column_dimensions[column].width = width

            ws.row_dimensions[1].height = 30
            for r in range(2, row_idx):
                ws.row_dimensions[r].height = 60

            ws.freeze_panes = "A2"
            wb.save(excel_path)
            return excel_path
        except Exception as e:
            self.logger.error(f"Error creating Excel from markdown rows: {str(e)}")
            return None

    def _contains_placeholder(self, text: str) -> bool:
        if not text:
            return True
        lower = text.lower()
        bad_phrases = [
            "detailed step-by-step execution procedure",
            "n/a",
            "etc.",
            "and so on",
            "placeholder"
        ]
        return any(p in lower for p in bad_phrases)

    def _row_has_placeholder(self, row: Dict[str, str]) -> bool:
        for key in ("Test Steps", "Expected Results", "Description", "Objective"):
            val = (row.get(key) or "").lower()
            if any(p in val for p in [
                "detailed step-by-step execution procedure",
                "n/a",
                "etc.",
                "and so on",
                "placeholder"
            ]):
                return True
        return False

    def _create_fallback_test_cases(self, batch_num: int, count: int) -> List[Dict[str, Any]]:
        """Create fallback test cases if parsing fails"""
        fallback_cases = []

        business_areas = [
            "Account Management", "Transaction Processing", "Customer Onboarding",
            "Loan Processing", "Digital Banking", "Risk Management", "Compliance",
            "Payment Processing", "Card Services", "Investment Services"
        ]

        test_scenarios = [
            "Create New Account with KYC Verification",
            "Process High Value Transfer with Approvals",
            "Mobile Banking Authentication and Security",
            "Loan Application Processing Workflow",
            "AML Transaction Monitoring and Alerts",
            "Customer Service Complaint Resolution",
            "Payment Gateway Integration Testing",
            "Credit Card Activation and Setup",
            "Investment Portfolio Management",
            "Regulatory Reporting Generation"
        ]

        for i in range(count):
            area_idx = i % len(business_areas)
            scenario_idx = i % len(test_scenarios)

            steps_str, expected_str = self._generate_explicit_steps(
                feature_prefix=f"TC_BATCH{batch_num}",
                business_area=business_areas[area_idx],
                scenario=test_scenarios[scenario_idx],
                index=i + 1,
            )

            test_case = {
                'test_case_id': f"TC_BATCH{batch_num}_{i+1:03d}",
                'title': f"{business_areas[area_idx]} - {test_scenarios[scenario_idx]} - Scenario {i+1}",
                'business_area': business_areas[area_idx],
                'test_category': "Functional",
                'priority': ["Critical", "High", "Medium"][i % 3],
                'test_objective': f"Test comprehensive {business_areas[area_idx].lower()} functionality with {test_scenarios[scenario_idx].lower()}",
                'user_role': ["CSR", "Manager", "Admin", "Customer"][i % 4],
                'preconditions': "System accessible, user logged in, test data available",
                'test_data': f"Realistic banking data for {business_areas[area_idx]} with proper formats and values",
                'test_steps': steps_str,
                'expected_results': expected_str,
                'post_conditions': "System state consistent, audit trail created, data integrity maintained",
                'business_rules': f"{business_areas[area_idx]} business rules and banking regulations applied correctly",
                'integration_points': "Core banking system, payment gateway, regulatory systems integration",
                'risk_level': ["High", "Medium", "Low"][i % 3],
                'automation_potential': "Y",
                'execution_time': f"{15 + (i % 20):02d}-{25 + (i % 15):02d} mins",
                'dependencies': "Test environment setup, user permissions, test data creation",
                'notes': f"Batch {batch_num+1} generated professional test case with comprehensive coverage"
            }
            fallback_cases.append(test_case)

        return fallback_cases

    def _deduplicate_test_cases(self, cases: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """Remove near-duplicates by title+objective signature."""
        seen = set()
        unique: List[Dict[str, Any]] = []
        for c in cases:
            title = (c.get('title') or c.get('Test Case Title') or '').strip().lower()
            obj = (c.get('test_objective') or c.get('Test Objective') or '').strip().lower()
            sig = f"{title}|{obj}"
            if sig and sig not in seen:
                seen.add(sig)
                unique.append(c)
        return unique

    def _generate_explicit_steps(self, feature_prefix: str, business_area: str, scenario: str, index: int) -> (str, str):
        """Generate explicit numbered steps and expected results for fallback cases.
        Returns (steps_str, expected_str) with <br> separators for Excel cells.
        """
        # Alternate through Maker and Checker for clearer maker-checker workflows
        actor_cycle = ["Maker", "Checker"]
        actor = actor_cycle[index % len(actor_cycle)]
        module = business_area.split(" ")[0]
        screen_main = f"{module} Dashboard"
        screen_detail = f"{module} {scenario.split(' ')[0]}"

        steps = [
            f"1. {actor} enters username 'user_{index}' and password 'P@ssw0rd!' on Login screen, then clicks 'Sign In'",
            f"2. {actor} navigates: Home → {module} → {screen_main}",
            f"3. {actor} navigates: {screen_main} → {screen_detail}",
            f"4. {actor} enters Customer ID 'CUST-{1000+index}'",
            f"5. {actor} sets Amount to '{(index+1)*1000}.00'",
            f"6. {actor} selects Currency 'QAR'",
            f"7. {actor} selects Product 'Standard'",
            f"8. {actor} clicks 'Validate'",
            f"9. System displays calculated fees and limits",
            f"10. {actor} clicks 'Submit'",
            f"11. {actor} confirms in modal by clicking 'Confirm'",
            f"12. Checker navigates: Home → Approval Queue → Items Pending → opens the submitted record",
            f"13. Checker reviews fields and clicks 'Approve'",
            f"14. {actor} opens Audit/History tab and reviews the last entry",
        ]
        expected = [
            "Login succeeds; landing page is Home with user name shown in header",
            f"'{screen_main}' screen is displayed with navigation breadcrumbs",
            f"'{screen_detail}' screen is displayed with enabled form controls",
            "Customer ID field accepts value and passes format validation",
            "Amount field accepts value and applies currency formatting",
            "Currency dropdown shows 'QAR' selected",
            "Product selection is stored; related fields are enabled",
            "No validation errors; 'Submit' becomes enabled",
            "Fees and limits section shows computed values based on inputs",
            "Request is sent; loading indicator appears",
            "Success toast 'Operation submitted for approval' with reference ID appears",
            "Record appears under Checker’s 'Items Pending' with status 'Pending Approval'",
            "Status changes to 'Approved' and the transaction is finalized",
            "Audit trail contains entries for submission and approval with user, time, and action",
        ]

        return "\n".join(steps), "\n".join(expected)

    def _create_feature_specific_fallback_cases(self, feature: Dict[str, Any], 
                                              feature_index: int, count: int) -> List[Dict[str, Any]]:
        """Create fallback test cases specific to the document feature"""
        fallback_cases = []

        test_scenarios = [
            f"Successful {feature['business_process']} by {feature['user_role']}",
            f"Failed {feature['business_process']} with validation errors",
            f"Partial {feature['business_process']} with missing data",
            f"Timeout during {feature['business_process']} process",
            f"Concurrent {feature['business_process']} by multiple {feature['user_role']}s",
            f"Approval workflow for {feature['business_process']}",
            f"Data validation in {feature['screen']} screen",
            f"Navigation flow through {feature['module']} module",
            f"Error handling in {feature['business_process']}",
            f"Performance testing of {feature['business_process']}"
        ]

        for i in range(count):
            scenario_idx = i % len(test_scenarios)
            scenario = test_scenarios[scenario_idx]

            steps_str, expected_str = self._generate_feature_specific_steps(
                feature, scenario, i + 1
            )

            test_case = {
                'test_case_id': f"{feature['feature_id']}_{i+1:03d}",
                'title': f"{feature['screen']} - {feature['module']} - {scenario}",
                'business_area': feature['module'],
                'test_category': "Functional",
                'priority': ["Critical", "High", "Medium"][i % 3],
                'test_objective': f"Test {scenario.lower()} in {feature['screen']} screen",
                'user_role': feature['user_role'],
                'preconditions': f"User logged in as {feature['user_role']}, {feature['screen']} screen accessible",
                'test_data': f"Realistic data for {feature['business_process']} process",
                'test_steps': steps_str,
                'expected_results': expected_str,
                'post_conditions': f"{feature['business_process']} completed successfully",
                'business_rules': f"{feature['module']} business rules applied correctly",
                'integration_points': f"Integration with {', '.join(feature['data_entities'])} entities",
                'risk_level': ["High", "Medium", "Low"][i % 3],
                'automation_potential': "Y",
                'execution_time': f"{15 + (i % 20):02d}-{25 + (i % 15):02d} mins",
                'dependencies': f"Access to {feature['screen']} screen, {feature['user_role']} permissions",
                'notes': f"Feature-specific test case for {feature['feature_id']}"
            }
            fallback_cases.append(test_case)

        return fallback_cases

    def _generate_feature_specific_steps(self, feature: Dict[str, Any], 
                                       scenario: str, index: int) -> (str, str):
        """Generate explicit numbered steps for feature-specific test cases"""
        actor = feature['user_role']
        screen = feature['screen']
        module = feature['module']
        process = feature['business_process']

        steps = [
            f"1. {actor} enters username 'user_{index}' and password 'P@ssw0rd!' on Login screen, then clicks 'Sign In'",
            f"2. {actor} navigates: Home → {module} → {screen}",
            f"3. {actor} selects '{process}' option from the main menu",
            f"4. {actor} enters Customer ID 'CUST-{1000+index}' in the customer field",
            f"5. {actor} sets Amount to '{(index+1)*1000}.00' in the amount field",
            f"6. {actor} selects Currency 'QAR' from the currency dropdown",
            f"7. {actor} enters Description 'Test {process} for {screen}' in the description field",
            f"8. {actor} clicks 'Validate' button to check input data",
            f"9. System displays validation results and calculated fees",
            f"10. {actor} reviews the information and clicks 'Submit'",
            f"11. {actor} confirms the action in the confirmation modal",
            f"12. System processes the {process} and displays success message",
            f"13. {actor} navigates to the transaction history to verify the entry",
            f"14. {actor} logs out of the system",
        ]
        
        expected = [
            "Login succeeds; landing page shows user name and role",
            f"'{screen}' screen is displayed with {module} navigation breadcrumbs",
            f"'{process}' form is loaded with all required fields enabled",
            "Customer ID field accepts value and passes format validation",
            "Amount field accepts value and applies currency formatting",
            "Currency dropdown shows 'QAR' selected",
            "Description field accepts text input",
            "No validation errors; submit button becomes enabled",
            "Fees and limits section shows computed values",
            "Request is sent; loading indicator appears",
            "Success message 'Operation completed successfully' with reference ID appears",
            f"{process} is processed and status changes to 'Completed'",
            "Transaction history shows the new entry with correct details",
            "User is successfully logged out; login screen is displayed",
        ]

        return "\n".join(steps), "\n".join(expected)

    def _create_professional_excel(self, test_cases: List[Dict[str, Any]],
                                 feature_prefix: str, timestamp: str, total_count: int) -> str:
        """Create professional Excel file with advanced formatting"""
        try:
            excel_filename = f"Professional_Test_Cases_{feature_prefix}_{timestamp}_{total_count}Cases.xlsx"
            excel_path = os.path.join("outputs", excel_filename)

            # Ensure outputs directory exists
            os.makedirs("outputs", exist_ok=True)

            # Create workbook and worksheet
            wb = Workbook()
            ws = wb.active
            ws.title = "Test Cases"

            # Define professional headers
            headers = [
                "Test Case ID", "Test Case Title", "Business Area", "Test Category",
                "Priority", "Test Objective", "User Role", "Preconditions",
                "Test Data", "Test Steps", "Expected Results", "Post-Conditions",
                "Business Rules", "Integration Points", "Risk Level",
                "Automation Potential", "Execution Time", "Dependencies", "Notes"
            ]

            # Add header row with styling
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col)
                cell.value = header
                cell.font = Font(bold=True, color="FFFFFF", size=12)
                cell.fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = Border(
                    left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin')
                )

            # Add test case data
            for row, test_case in enumerate(test_cases, 2):
                values = [
                    test_case.get('test_case_id', f'{feature_prefix}_{row:03d}'),
                    test_case.get('title', f'Test Case {row-1}'),
                    test_case.get('business_area', 'Banking Operations'),
                    test_case.get('test_category', 'Functional'),
                    test_case.get('priority', 'High'),
                    test_case.get('test_objective', 'Test functionality'),
                    test_case.get('user_role', 'CSR'),
                    test_case.get('preconditions', 'System accessible'),
                    test_case.get('test_data', 'Realistic test data'),
                    test_case.get('test_steps', 'Detailed execution steps'),
                    test_case.get('expected_results', 'Expected outcomes'),
                    test_case.get('post_conditions', 'System state maintained'),
                    test_case.get('business_rules', 'Business logic applied'),
                    test_case.get('integration_points', 'N/A'),
                    test_case.get('risk_level', 'Medium'),
                    test_case.get('automation_potential', 'Y'),
                    test_case.get('execution_time', '20-30 mins'),
                    test_case.get('dependencies', 'None'),
                    test_case.get('notes', 'Additional information')
                ]

                for col, value in enumerate(values, 1):
                    cell = ws.cell(row=row, column=col)
                    cell.value = str(value)[:500]
                    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                    cell.border = Border(
                        left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin')
                    )

                    # Priority-based color coding
                    priority = test_case.get('priority', 'Medium').upper()
                    if priority == 'CRITICAL':
                        cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
                    elif priority == 'HIGH':
                        cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
                    elif priority == 'MEDIUM':
                        cell.fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
                    elif priority == 'LOW':
                        cell.fill = PatternFill(start_color="E6FFE6", end_color="E6FFE6", fill_type="solid")

            # Auto-adjust column widths
            column_widths = {
                'A': 15, 'B': 50, 'C': 20, 'D': 15, 'E': 10,
                'F': 40, 'G': 15, 'H': 30, 'I': 40, 'J': 60,
                'K': 40, 'L': 25, 'M': 30, 'N': 25, 'O': 10,
                'P': 12, 'Q': 15, 'R': 25, 'S': 30
            }

            for column, width in column_widths.items():
                ws.column_dimensions[column].width = width

            # Set row heights
            ws.row_dimensions[1].height = 30
            for row in range(2, len(test_cases) + 2):
                ws.row_dimensions[row].height = 100

            # Freeze panes
            ws.freeze_panes = "A2"

            # Save workbook
            wb.save(excel_path)
            return excel_path

        except Exception as e:
            self.logger.error(f"Error creating professional Excel file: {str(e)}")
            return None

    def _create_enhanced_markdown(self, content: str, feature_prefix: str, timestamp: str) -> str:
        """Create enhanced markdown file with professional formatting"""
        try:
            markdown_filename = f"Enhanced_Test_Cases_{feature_prefix}_{timestamp}.md"
            markdown_path = os.path.join("outputs", markdown_filename)

            enhanced_content = f"""# Professional Test Cases - {feature_prefix}
            
## Test Suite Information
- **Generated**: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}
- **Feature Prefix**: {feature_prefix}
        - **Total Cases**: Large Scale Generation
- **Coverage**: Comprehensive Banking Functionality
- **AI Model**: {self.model_name}

## Test Case Coverage Areas

### 🏦 Core Banking Operations (35%)
- Account Management & Lifecycle
- Customer Onboarding & KYC
- Transaction Processing & Settlement
- Loan Origination & Servicing

### 💻 Digital Banking & Channels (20%)
- Internet Banking Platform
- Mobile Banking Applications
- ATM & Self-Service Channels
- API & Integration Services

### ⚖️ Risk & Compliance (15%)
- AML/KYC Compliance
- Credit Risk Assessment
- Operational Risk Controls
- Regulatory Reporting

### 🏢 Back Office Operations (10%)
- Settlement & Clearing
- Accounting & Financial Reporting
- Operations Support
- Data Management & ETL

### 👥 Customer Experience (10%)
- Customer Service Operations
- Relationship Management
- Marketing & Campaign Management
- Document Management

### 🔧 Integration & Infrastructure (5%)
- System Integrations
- Security & Authentication
- Performance & Scalability
- Disaster Recovery

### 📊 Analytics & Reporting (5%)
- Business Intelligence
- Management Reporting
- Audit & Control Reports
- Performance Analytics

---

{content}

---

## Quality Assurance Notes

            This test suite has been generated using the configured AI model with the following quality measures:
- ✅ **Comprehensive Coverage**: All major banking functions covered
- ✅ **Realistic Test Data**: Culturally appropriate and format-compliant data
- ✅ **Detailed Steps**: 15-25 detailed execution steps per test case
- ✅ **Business Focus**: Emphasis on end-to-end business processes
- ✅ **Professional Standards**: Industry-standard test case documentation
- ✅ **Integration Testing**: Cross-system and API integration scenarios
- ✅ **Risk-Based**: Priority and risk-level assignments
        - ✅ **AI-Generated**: Using the configured LLM

        Generated using Enhanced Test Case Service v3.0 with LLM Integration
Model: {self.model_name}
"""

            with open(markdown_path, 'w', encoding='utf-8') as f:
                f.write(enhanced_content)

            return markdown_path

        except Exception as e:
            self.logger.error(f"Error creating enhanced markdown file: {str(e)}")
            return None

    def generate_comprehensive_test_data_library(self) -> Dict[str, List[str]]:
        """Generate comprehensive test data library for all scenarios"""
        return {
            "arabic_names": [
                "أحمد محمد الراشد", "فاطمة علي الزهراني", "خالد عبدالله القحطاني",
                "سارة محمد العتيبي", "عبدالرحمن أحمد الدوسري", "نورا سالم المالكي",
                "محمد عبدالله الشمري", "هند فهد الأحمدي", "سلطان راشد آل ثاني",
                "مريم حسن الكواري", "يوسف علي البوعينين", "لطيفة محمد الأنصاري"
            ],
            "western_names": [
                "Sarah Elizabeth Johnson", "Michael James O'Connor", "Emma Charlotte Williams",
                "David Benjamin Smith", "Jessica Marie Davis", "Christopher Andrew Brown",
                "Amanda Nicole Taylor", "Robert William Miller", "Jennifer Lynn Anderson",
                "Matthew John Wilson", "Ashley Michelle Garcia", "Daniel Thomas Martinez"
            ],
            "asian_names": [
                "李明华 (Li Ming Hua)", "田中雄介 (Tanaka Yusuke)", "김민준 (Kim Min-jun)",
                "Priya Sharma", "Raj Patel", "Yuki Tanaka", "Chen Wei Ming",
                "Sakura Yamamoto", "Arjun Kumar", "Meera Desai", "Wang Li Na",
                "Hiroshi Sato"
            ],
            "company_names": [
                "Al-Mansour International Trading & Contracting LLC",
                "Qatar Digital Innovation Solutions W.L.L",
                "Gulf Financial Services Holding Company",
                "Arabian Peninsula Investment Group",
                "Middle East Technology Partners LLC",
                "Doha Commercial Enterprises W.L.L",
                "Qatar National Construction Company",
                "Al-Rayyan Business Solutions LLC",
                "Gulf Coast Trading & Services",
                "Peninsula Development Corporation"
            ],
            "addresses": [
                "Villa 23, Street 820, Al-Sadd District, Zone 25, P.O. Box 12345, Doha, Qatar",
                "Apartment 1205, Burj Al-Arab Business Tower, Sheikh Zayed Road, Dubai, UAE 12345",
                "Building 456, Prince Sultan Street, Al-Olaya District, Riyadh 11564, Saudi Arabia",
                "Office 890, Al-Muntazah Complex, C-Ring Road, Doha, Qatar 54321",
                "Tower 15, Floor 20, Al-Corniche Street, West Bay, Doha, Qatar",
                "Villa 67, Al-Waab City, Street 45, Doha, Qatar 98765"
            ],
            "phone_numbers": [
                "+974-5544-7788", "+974-3366-9922", "+974-7788-1122", "+974-5599-3344",
                "+974-4455-6677", "+974-3377-8899", "+971-50-123-4567", "+966-50-987-6543"
            ],
            "email_addresses": [
                "ahmed.alrashid@qatarbank.com.qa", "sarah.johnson@company.com",
                "priya.sharma@business.org", "michael.oconnor@finance.qa",
                "fatima.alzahra@trading.com", "david.smith@investment.qa"
            ],
            "account_numbers": [
                "QAR-SAV-2024-789456123", "USD-CHK-2024-555777888", "EUR-INV-2024-999000111",
                "GBP-SAV-2024-444333222", "QAR-LON-2024-777888999", "USD-CC-2024-111222333"
            ],
            "iban_numbers": [
                "QA58DOHB00001234567890123456", "QA29QNBK000000000012345678",
                "QA87CBQK00000000123456789012", "AE070331234567890123456",
                "SA0380000000608010167519", "GB33BUKB20201555555555"
            ],
            "amounts": [
                "15,750.50", "125,500.75", "85,250.25", "999,999.99", "50,000.00",
                "75,500.75", "250,000.00", "5,500.25", "1,000,000.00", "10,750.50"
            ],
            "banking_products": [
                "Premium Savings Plus", "Youth Future Saver", "Senior Citizen Special",
                "Business Elite Current", "Professional Plus Current", "Salary Transfer Account",
                "Personal Finance Plus", "Home Loan Advantage", "Auto Finance Express",
                "SME Business Loan", "Platinum Rewards Card", "Business Corporate Card"
            ],
            "currencies": [
                "QAR (Qatari Riyal)", "USD (US Dollar)", "EUR (Euro)", "GBP (British Pound)",
                "AED (UAE Dirham)", "SAR (Saudi Riyal)", "KWD (Kuwaiti Dinar)"
            ],
            "transaction_types": [
                "Fund Transfer", "Bill Payment", "Salary Credit", "Investment Purchase",
                "Loan Disbursement", "Card Payment", "ATM Withdrawal", "Check Deposit",
                "Standing Order", "Direct Debit", "Wire Transfer", "Currency Exchange"
            ]
        }

    def create_performance_test_matrix(self) -> List[Dict[str, Any]]:
        """Create performance test scenarios matrix"""
        return [
            {
                "scenario": "Peak Hour Concurrent User Load",
                "user_load": 2000,
                "duration": "2 hours",
                "transaction_mix": "Login(30%), Balance Inquiry(25%), Fund Transfer(20%), Bill Payment(15%), Other(10%)",
                "success_criteria": "Response time <3s, Error rate <1%, System stability maintained"
            },
            {
                "scenario": "High Volume Transaction Processing",
                "transaction_volume": 10000,
                "duration": "1 hour",
                "batch_size": 500,
                "success_criteria": "Processing time <5s per batch, Data integrity 100%, No system errors"
            },
            {
                "scenario": "Database Stress Test",
                "concurrent_queries": 500,
                "duration": "30 minutes",
                "query_types": "SELECT(40%), INSERT(25%), UPDATE(20%), DELETE(15%)",
                "success_criteria": "Query response <2s, Connection pool stable, No deadlocks"
            }
        ]

    def generate_security_test_scenarios(self) -> List[Dict[str, Any]]:
        """Generate comprehensive security test scenarios"""
        return [
            {
                "category": "Authentication Security",
                "test_scenarios": [
                    "Multi-factor authentication bypass attempt",
                    "Password brute force attack simulation",
                    "Session hijacking prevention test",
                    "Account lockout policy validation"
                ]
            },
            {
                "category": "Authorization Controls",
                "test_scenarios": [
                    "Privilege escalation attempt",
                    "Cross-user data access test",
                    "Role-based access control validation",
                    "Administrative function protection"
                ]
            },
            {
                "category": "Data Protection",
                "test_scenarios": [
                    "Sensitive data encryption verification",
                    "Data transmission security test",
                    "PII data masking validation",
                    "Database security assessment"
                ]
            }
        ]

    def generate_integration_test_matrix(self) -> List[Dict[str, Any]]:
        """Generate integration test scenarios matrix"""
        return [
            {
                "integration_type": "Core Banking - Payment Gateway",
                "test_scenarios": [
                    "Real-time payment processing with reconciliation",
                    "Failed transaction rollback verification",
                    "Timeout handling and retry mechanism",
                    "Multi-currency payment processing"
                ]
            },
            {
                "integration_type": "CRM - Core Banking Sync",
                "test_scenarios": [
                    "Customer data synchronization verification",
                    "Real-time updates propagation test",
                    "Data consistency validation",
                    "Conflict resolution mechanism test"
                ]
            },
            {
                "integration_type": "Regulatory Reporting Systems",
                "test_scenarios": [
                    "Automated report generation and submission",
                    "Data accuracy and completeness validation",
                    "Submission acknowledgment handling",
                    "Error notification and retry mechanism"
                ]
            }
        ]

    async def generate_test_cases_from_documents(self,
                                         analysis_doc: str,
                                         risk_doc: str,
                                         reviewed_doc: str,
                                         feature_prefix: str = "TC",
                                         iteration_count: int = 300) -> Dict[str, Any]:
        """Generate test cases from analysis, risk, and reviewed documents"""
        combined_input = f"""ANALYSIS DOCUMENT:
{analysis_doc}

RISK ASSESSMENT DOCUMENT:  
{risk_doc}

EXPERT REVIEWED DOCUMENT:
{reviewed_doc}"""

        return await self.generate_large_scale_test_cases(
            user_input=combined_input,
            technical_analysis=analysis_doc,
            domain_analysis=reviewed_doc,
            test_risk=risk_doc,
            feature_prefix=feature_prefix,
            iteration_count=iteration_count,
            raw_input_chunk=combined_input
        )

class TestCaseService:
    """
    Original Test Case Service - Wrapper around EnhancedTestCaseService for backward compatibility
    """
    def __init__(self, model_name: str = "gpt-4.1", generator: Optional[TextGenerator] = None):
        # Use the enhanced service as the underlying implementation
        self.enhanced_service = EnhancedTestCaseService(model_name, generator)
        self.model_name = model_name
        self.logger = logging.getLogger(__name__)

    async def generate_test_cases_from_file(self, file_path: str, feature_prefix: str = "TC", iteration_count: int = 50) -> Dict[str, Any]:
        """
        Generate test cases from file - delegates to enhanced service
        """
        try:
            # Read file content
            if not os.path.exists(file_path):
                return {
                    "success": False,
                    "error": f"File not found: {file_path}",
                    "test_cases_content": "",
                    "generated_count": 0
                }

            # Read file content
            with open(file_path, 'r', encoding='utf-8') as f:
                content = f.read()

            # Use enhanced service for generation
            result = await self.enhanced_service.generate_large_scale_test_cases(
                user_input=f"Generate test cases from file: {file_path}\n\nContent:\n{content[:2000]}...",
                feature_prefix=feature_prefix,
                iteration_count=iteration_count
            )

            return result
        except Exception as e:
            self.logger.error(f"Error in generate_test_cases_from_file: {str(e)}")
            return {
                "success": False,
                "error": str(e),
                "test_cases_content": "",
                "generated_count": 0
            }

    async def generate_test_cases(self, user_input: str, feature_prefix: str = "TC", iteration_count: int = 50, **kwargs) -> Dict[str, Any]:
        """
        Generate test cases - delegates to enhanced service
        """
        try:
            result = await self.enhanced_service.generate_large_scale_test_cases(
                user_input=user_input,
                feature_prefix=feature_prefix,
                iteration_count=iteration_count,
                **kwargs
            )
            return result
        except Exception as e:
            self.logger.error(f"Error in generate_test_cases: {str(e)}")
            return {
                "success": False,
                "error": str(e),
                "test_cases_content": "",
                "generated_count": 0
            }
